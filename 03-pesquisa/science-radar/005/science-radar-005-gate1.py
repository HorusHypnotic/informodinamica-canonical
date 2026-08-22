from pathlib import Path
from collections import defaultdict
import csv, re, json
from datetime import datetime
from openpyxl import load_workbook

base=Path('/home/ubuntu/science-radar-004-recovery/plos/s1-extracted/data/team-data')
logdir=base/'team-chat-logs'
wb=load_workbook(base/'team-data.xlsx', read_only=True, data_only=True)
ws=wb.active
rows=list(ws.iter_rows(values_only=True))
headers=[str(x) if x is not None else '' for x in rows[0]]
hidx={h:i for i,h in enumerate(headers)}
work={str(r[hidx['Session']]): r for r in rows[1:] if r[hidx['Session']] is not None}

logs=defaultdict(list)
for fp in sorted(logdir.glob('*.csv')):
    with fp.open(encoding='utf-8-sig', newline='', errors='replace') as f:
        for row in csv.DictReader(f):
            s=row.get('session','').strip()
            if not s: continue
            ts=row.get('timestamp','').strip()
            try: dt=datetime.fromisoformat(ts)
            except: dt=None
            row['_dt']=dt
            logs[s].append(row)

def task_norm(x):
    x=x.strip()
    x=re.sub(r'^XVal Session [0-9]+-', '', x)
    x=re.sub(r'-[0-9]+$', '', x)
    return x.strip()

report=[]
for s, events in sorted(logs.items()):
    events=[e for e in events if e['_dt'] is not None]
    starts=[]
    seen=set()
    for e in sorted(events,key=lambda z:z['_dt']):
        ct=e.get('completed_task','').strip()
        if not ct: continue
        t=task_norm(ct)
        if e.get('event')=='Load Instructions' and t not in seen:
            starts.append((t,e['_dt']))
            seen.add(t)
    target=[x for x in starts if x[0]=='Matrix Solving']
    if not target: continue
    t0,tc=target[0]
    idx=starts.index(target[0])
    nxt=starts[idx+1] if idx+1<len(starts) else None
    score=work.get(s,[None]*len(headers))[hidx.get('Matrix Solving',-1)] if s in work else None
    report.append({'session':s,'in_workbook':s in work,'target_start':tc.isoformat(),'next_task':nxt[0] if nxt else None,'ty_proxy':nxt[1].isoformat() if nxt else None,'tc_lt_ty':bool(nxt and tc<nxt[1]),'score':score,'n_starts':len(starts),'starts':[x[0] for x in starts]})

matches=[x for x in report if x['in_workbook']]
valid=[x for x in matches if x['score'] is not None and x['ty_proxy'] and x['tc_lt_ty']]
print('WORKBOOK_SESSIONS',len(work))
print('LOG_SESSIONS',len(logs))
print('MATRIX_TARGET_SESSIONS',len(report))
print('MATRIX_WITH_WORKBOOK',len(matches))
print('VALID_SCORE_TC_TY',len(valid))
print('NO_WORKBOOK',sum(not x['in_workbook'] for x in report))
print('NO_NEXT_TASK',sum(x['next_task'] is None for x in matches))
print('SCORE_MISSING',sum(x['score'] is None for x in matches))
print('NEXT_TASKS',sorted({x['next_task'] for x in matches if x['next_task']}))
print('VALID_SAMPLE')
for x in valid[:10]: print(x)
Path('/home/ubuntu/science-radar-005-gate1.json').write_text(json.dumps({'workbook_sessions':len(work),'log_sessions':len(logs),'matrix_target_sessions':len(report),'matrix_with_workbook':len(matches),'valid_score_tc_ty':len(valid),'records':report},indent=2,ensure_ascii=False),encoding='utf-8')

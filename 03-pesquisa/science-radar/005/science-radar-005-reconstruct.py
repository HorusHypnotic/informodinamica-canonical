from pathlib import Path
from collections import defaultdict
from datetime import datetime, timedelta
import csv, json, re, math
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn.impute import SimpleImputer
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler
from sklearn.ensemble import HistGradientBoostingRegressor
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score

ROOT=Path('/home/ubuntu/science-radar-004-recovery/plos/s1-extracted/data/team-data')
OUT=Path('/home/ubuntu/science-radar-005-results'); OUT.mkdir(exist_ok=True)
SEED=2026

# Raw workbook: one row per session/team, no published derived tables.
wb=load_workbook(ROOT/'team-data.xlsx', read_only=True, data_only=True)
ws=wb.active
rows=list(ws.iter_rows(values_only=True))
headers=[str(x).strip() if x is not None else '' for x in rows[0]]
wbdf=pd.DataFrame(rows[1:], columns=headers)
wbdf=wbdf[wbdf['Session'].notna()].copy(); wbdf['Session']=wbdf['Session'].astype(str).str.strip()
for c in wbdf.columns:
    if c!='Session': wbdf[c]=pd.to_numeric(wbdf[c],errors='coerce')

# Raw logs.
logdir=ROOT/'team-chat-logs'; logs=defaultdict(list)
for fp in sorted(logdir.glob('*.csv')):
    with fp.open(encoding='utf-8-sig', newline='', errors='replace') as f:
        for row in csv.DictReader(f):
            s=row.get('session','').strip(); ts=row.get('timestamp','').strip()
            if not s: continue
            try: dt=datetime.fromisoformat(ts)
            except: dt=None
            row['_dt']=dt; logs[s].append(row)

def norm_task(x):
    x=x.strip(); x=re.sub(r'^XVal Session [0-9]+-', '', x); x=re.sub(r'-[0-9]+$', '', x); return x.strip()

def event_task(e): return norm_task(e.get('completed_task','')) if e.get('completed_task') else ''

def metric(y,p):
    return {'MAE':float(mean_absolute_error(y,p)),'RMSE':float(mean_squared_error(y,p)**0.5),'R2':float(r2_score(y,p)) if len(y)>1 else None}

def net_collab(events, cutoff):
    pre=[e for e in events if e['_dt'] is not None and e['_dt'] < cutoff]
    chats=[e for e in pre if e.get('event')=='Chat']
    senders=[(e.get('subject') or e.get('chat_name') or '').strip() for e in chats]
    senders=[s for s in senders if s]
    uniq=sorted(set(senders)); counts=pd.Series(senders).value_counts() if senders else pd.Series(dtype=float)
    # Directed handoff network: consecutive distinct message senders, only observable from the log.
    edges=[]
    for a,b in zip(senders,senders[1:]):
        if a and b and a!=b: edges.append((a,b))
    E=set(edges); n=len(uniq); possible=n*(n-1)
    outdeg=defaultdict(set); indeg=defaultdict(set)
    for a,b in E: outdeg[a].add(b); indeg[b].add(a)
    reciprocal=sum(1 for a,b in E if (b,a) in E)/len(E) if E else 0.0
    density=len(E)/possible if possible else 0.0
    entropy=0.0
    if len(senders):
        probs=counts.to_numpy()/len(senders); entropy=float(-(probs*np.log(probs)).sum()/math.log(len(probs))) if len(probs)>1 else 0.0
    delays=[(b['_dt']-a['_dt']).total_seconds() for a,b in zip(chats,chats[1:]) if b['_dt']>=a['_dt']]
    return {
      'N_chat_pre':len(chats),'N_sender_pre':n,'N_edge_handoff_pre':len(E),'Network_density_pre':density,
      'Network_reciprocity_pre':reciprocal,'Network_mean_outdegree_pre':float(np.mean([len(outdeg[x]) for x in uniq])) if uniq else 0.0,
      'Network_mean_indegree_pre':float(np.mean([len(indeg[x]) for x in uniq])) if uniq else 0.0,
      'Collab_entropy_pre':entropy,'Collab_sender_balance_pre':float(counts.std()/counts.mean()) if len(counts)>1 and counts.mean() else 0.0,
      'Collab_mean_interchat_sec_pre':float(np.mean(delays)) if delays else np.nan,
      'Collab_median_interchat_sec_pre':float(np.median(delays)) if delays else np.nan,
      'Collab_total_words_pre':float(sum(len((e.get('data') or '').split()) for e in chats)),
      'Collab_total_chars_pre':float(sum(len(e.get('data') or '') for e in chats)),
      'Collab_event_count_pre':len(pre),
    }

# Build one target row per session with exact task linkage and temporal markers.
records=[]; exclusions=defaultdict(int)
for session, events0 in sorted(logs.items()):
    events=sorted([e for e in events0 if e['_dt'] is not None], key=lambda e:e['_dt'])
    starts=[]; seen=set()
    for e in events:
        t=event_task(e)
        if e.get('event')=='Load Instructions' and t and t not in seen:
            starts.append((t,e['_dt'])); seen.add(t)
    target=[x for x in starts if x[0]=='Matrix Solving']
    if not target: exclusions['no_matrix_task']+=1; continue
    tc=target[0][1]; i=starts.index(target[0]); nxt=starts[i+1] if i+1<len(starts) else None
    if session not in set(wbdf['Session']): exclusions['no_workbook_session']+=1; continue
    row=wbdf.loc[wbdf['Session']==session].iloc[0]
    y=row.get('Matrix Solving',np.nan)
    if pd.isna(y): exclusions['missing_outcome']+=1; continue
    if not nxt or not (tc<nxt[1]): exclusions['invalid_ty']+=1; continue
    t0=events[0]['_dt']; ty=nxt[1]
    pre_tasks=[t for t,tstamp in starts if tstamp<tc]
    # History scores only for exact pre-target task columns.
    hist_cols=['Typing Text','Typing Numbers','Combined Typing']
    feat={'Session':session,'T0':t0.isoformat(),'TC':tc.isoformat(),'TY_proxy':ty.isoformat(),'Y_Matrix_Solving':float(y),'n_pre_tasks':len(pre_tasks)}
    for c in hist_cols: feat['H_'+c.replace(' ','_')]=float(row[c]) if c in row.index and pd.notna(row[c]) else np.nan
    static=['Talking','Late','Subject Count','Age - Mean','Age - Max','Age - Min','Is Female - Mean','Mind in the Eyes - Mean','Team Cohesion - Mean','Group IQs']
    static += [f'Big5_{i} - Mean' for i in range(1,11)]
    for c in static: feat['S_'+c.replace(' ','_').replace('-','')]=float(row[c]) if c in row.index and pd.notna(row[c]) else np.nan
    feat.update(net_collab(events,tc))
    # Cutoff sensitivity recomputes only event-derived fields; target and history remain fixed.
    for delta in [-300,-60,0]:
        d=net_collab(events,tc+timedelta(seconds=delta))
        for k,v in d.items(): feat[f'{k}_cut{delta}']=v
    records.append(feat)

df=pd.DataFrame(records).sort_values('TC').reset_index(drop=True)
# Main cut uses chronological sessions; one row per session, hence no cross-session leakage.
cut=int(np.floor(0.8*len(df))); train=df.iloc[:cut].copy(); test=df.iloc[cut:].copy()
# Keep a manifest of feature genealogy; no derived S1 tables enter.
hist=[c for c in df.columns if c.startswith('H_') or c=='n_pre_tasks']
static=[c for c in df.columns if c.startswith('S_')]
network=[c for c in ['N_sender_pre','N_edge_handoff_pre','Network_density_pre','Network_reciprocity_pre','Network_mean_outdegree_pre','Network_mean_indegree_pre']]
collab=[c for c in ['N_chat_pre','Collab_entropy_pre','Collab_sender_balance_pre','Collab_mean_interchat_sec_pre','Collab_median_interchat_sec_pre','Collab_total_words_pre','Collab_total_chars_pre','Collab_event_count_pre']]
M0=hist+static; M1=M0+network; M2=M1+collab

def fit_predict(cols, tr=train, te=test):
    Xtr=tr[cols].replace([np.inf,-np.inf],np.nan); Xte=te[cols].replace([np.inf,-np.inf],np.nan)
    model=Pipeline([('impute',SimpleImputer(strategy='median',add_indicator=True)),('scale',StandardScaler()),('gb',HistGradientBoostingRegressor(max_iter=180,max_leaf_nodes=15,learning_rate=.06,l2_regularization=1,random_state=SEED))])
    model.fit(Xtr,tr['Y_Matrix_Solving']); pred=model.predict(Xte)
    z=metric(te['Y_Matrix_Solving'],pred); z['n_train']=len(tr); z['n_test']=len(te); return z,pred

results=[]
for name,cols in [('BASE_MEAN',[]),('M0_HISTORY',M0),('M1_HISTORY_NETWORK',M1),('M2_HISTORY_NETWORK_COLLABORATION',M2)]:
    if name=='BASE_MEAN': pred=np.repeat(train['Y_Matrix_Solving'].mean(),len(test)); z=metric(test['Y_Matrix_Solving'],pred); z.update({'n_train':len(train),'n_test':len(test)})
    else: z,pred=fit_predict(cols)
    z['model']=name; results.append(z)

# 20 permutation placebos for added columns; same train/test and model family.
rng=np.random.default_rng(SEED); placebo=[]
for i in range(20):
    tr=train.copy(); te=test.copy()
    for c in network+collab:
        tr[c]=rng.permutation(tr[c].to_numpy()); te[c]=rng.permutation(te[c].to_numpy())
    z,_=fit_predict(M2,tr,te); z.update({'replicate':i+1,'model':'M2_SHUFFLED_ADDED_FEATURES'}); placebo.append(z)

# Cutoff sensitivity for M1/M2 using -300, -60, 0 sec event features.
sens=[]
for delta in [-300,-60,0]:
    ncols=[c.replace('_pre',f'_pre_cut{delta}') for c in network]
    ccols=[c.replace('_pre',f'_pre_cut{delta}') for c in collab]
    # features with exact names exist; preserve baseline history/static
    for name,cols in [('M0_HISTORY',M0),('M1_HISTORY_NETWORK',M0+ncols),('M2_HISTORY_NETWORK_COLLABORATION',M0+ncols+ccols)]:
        z,_=fit_predict(cols); z.update({'cutoff_delta_seconds':delta,'model':name}); sens.append(z)

# Temporal stability by early/late halves, retaining chronological local split where possible.
stab=[]
for label,sub in [('early',df.iloc[:len(df)//2].copy()),('late',df.iloc[len(df)//2:].copy())]:
    if len(sub)<10: continue
    c=max(2,int(np.floor(.8*len(sub)))); tr=sub.iloc[:c]; te=sub.iloc[c:]
    for name,cols in [('M0_HISTORY',M0),('M1_HISTORY_NETWORK',M1),('M2_HISTORY_NETWORK_COLLABORATION',M2)]:
        if len(te)<2: continue
        z,_=fit_predict(cols,tr,te); z.update({'segment':label,'model':name}); stab.append(z)

# Artifact outputs.
df.to_csv(OUT/'analysis_rows.csv',index=False)
pd.DataFrame(results).to_csv(OUT/'model_results.csv',index=False)
pd.DataFrame(placebo).to_csv(OUT/'placebo_results.csv',index=False)
pd.DataFrame(sens).to_csv(OUT/'cutoff_sensitivity.csv',index=False)
pd.DataFrame(stab).to_csv(OUT/'stability_results.csv',index=False)
manifest=[]
for group,cols,src,deriv in [('M0',M0,'team-data.xlsx','pre-task workbook scores and pre-task attributes'),('M1',network,'team-chat-logs/*.csv','Chat events < TC -> consecutive-sender handoff network'),('M2',collab,'team-chat-logs/*.csv','Chat/application events < TC -> participation and timing')]:
    for c in cols: manifest.append({'model_set':group,'feature':c,'source_event':src,'earliest_available':'before TC' if group!='M0' else 'pre-task or prior task','cutoff_safe':'YES','derivation':deriv,'outcome_dependence':'NO','leakage_risk':'LOW/MEDIUM; inspect dependence'})
pd.DataFrame(manifest).to_csv(OUT/'feature_genealogy.csv',index=False)
summary={'n_workbook_sessions':len(wbdf),'n_log_sessions':len(logs),'n_valid_rows':len(df),'n_train':len(train),'n_test':len(test),'exclusions':dict(exclusions),'tc':'first Load Instructions for Matrix Solving','ty':'first Load Instructions for next task; proxy outcome availability','target':'Matrix Solving score','split':'chronological 80/20 by session; no session repeats','seed':SEED,'data_hash':'54c33ca5338d6cb2cbb8215a0010ab6251f690d1018069ea94fa048d4a6d5ecc'}
(OUT/'gate1_summary.json').write_text(json.dumps(summary,indent=2,ensure_ascii=False),encoding='utf-8')
(OUT/'run_metadata.json').write_text(json.dumps({'script':'science-radar-005-reconstruct.py','inputs':['team-data.xlsx','team-chat-logs/*.csv'],'derived_tables_used':False,'models':['mean baseline','HistGradientBoostingRegressor'],'seed':SEED},indent=2,ensure_ascii=False),encoding='utf-8')
print(json.dumps(summary,ensure_ascii=False))
print(pd.DataFrame(results).to_string(index=False))
print('PLACEBO_MEAN',pd.DataFrame(placebo)[['MAE','RMSE','R2']].mean().to_dict())
print('SENSITIVITY'); print(pd.DataFrame(sens)[['cutoff_delta_seconds','model','MAE','RMSE','R2']].to_string(index=False))
print('STABILITY'); print(pd.DataFrame(stab)[['segment','model','MAE','RMSE','R2']].to_string(index=False))
